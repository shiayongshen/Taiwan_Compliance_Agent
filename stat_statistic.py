import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================
# 基本設定
# ============================================================
output_dir = Path("outputs")
excel_path = output_dir / "pipeline_statistics.xlsx"

# ============================================================
# 1️⃣ 收集所有統計資料
# ============================================================
all_stats = []

for filename in os.listdir(output_dir):
    if filename.endswith(".stats.json"):
        filepath = output_dir / filename
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        summary = {
            "case_id": data.get("case_id"),
            "success": data.get("success"),
            "error_message": data.get("error_message"),
            "total_time_sec": data.get("total_time_sec"),
            "repair_attempts": data.get("repair_attempts"),
            "total_agent_calls": data.get("total_agent_calls"),
            "total_input_tokens": data.get("total_input_tokens"),
            "total_output_tokens": data.get("total_output_tokens"),
            "total_tokens": data.get("total_tokens"),
            "total_cost_usd": data.get("total_cost_usd"),
            "timestamp": data.get("timestamp"),
        }

        checkpoints_detail = data.get("checkpoints_detail", {})

        def classify_status(v):
            if isinstance(v, dict):
                return "PASS" if v.get("passed") else "FAIL" if v.get("passed") is False else "NOT_RUN"
            elif isinstance(v, bool):
                return "PASS" if v else "FAIL"
            else:
                return "NOT_RUN"

        for key, value in checkpoints_detail.items():
            summary[key] = classify_status(value)

        # 🔧 修正邏輯：如果 step5_repair_needed 是 FAIL，則 step5_repair_success 標為 NOT_RUN
        if summary.get("step5_repair_needed") == "FAIL":
            summary["step5_repair_success"] = "NOT_RUN"

        all_stats.append(summary)

# ============================================================
# 2️⃣ 建立 DataFrame 並排序
# ============================================================
if not all_stats:
    raise ValueError("❌ 沒有找到任何 .stats.json 檔案")

all_stats.sort(key=lambda x: int(x["case_id"].split("_")[1]))
stats_df = pd.DataFrame(all_stats)

checkpoint_cols = [col for col in stats_df.columns if col.startswith("step")]
base_cols = [
    "case_id",
    "success",
    "total_time_sec",
    "repair_attempts",
    "total_agent_calls",
    "total_input_tokens",
    "total_output_tokens",
    "total_tokens",
    "total_cost_usd",
    "error_message",
    "timestamp",
]
stats_df = stats_df[base_cols + checkpoint_cols]

# ============================================================
# 3️⃣ 計算總體統計 (Overall)
# ============================================================
summary_data = {
    "Total Cases": len(all_stats),
    "Success Cases": stats_df["success"].sum(),
    "Failed Cases": (~stats_df["success"]).sum(),
    "Success Rate": f"{stats_df['success'].mean() * 100:.2f}%",
    "Avg Time (sec)": stats_df["total_time_sec"].mean(),
    "Avg Repair Attempts": stats_df["repair_attempts"].mean(),
    "Total Tokens": stats_df["total_tokens"].sum(),
    "Avg Tokens per Case": stats_df["total_tokens"].mean(),
    "Total Cost (USD)": stats_df["total_cost_usd"].sum(),
    "Avg Cost per Case (USD)": stats_df["total_cost_usd"].mean(),
}
summary_df = pd.DataFrame([summary_data])

# ============================================================
# 4️⃣ 檢查點通過率統計 (Checkpoints)
# ============================================================
checkpoint_stats = {}
for col in checkpoint_cols:
    pass_count = (stats_df[col] == "PASS").sum()
    fail_count = (stats_df[col] == "FAIL").sum()
    skip_count = (stats_df[col] == "SKIP").sum()
    not_run_count = (stats_df[col] == "NOT_RUN").sum()
    checkpoint_stats[col] = {
        "PASS": pass_count,
        "FAIL": fail_count,
        "SKIP": skip_count,
        "NOT_RUN": not_run_count,
        "Pass Rate": f"{(pass_count / len(stats_df) * 100):.2f}%",
    }

checkpoint_df = pd.DataFrame(checkpoint_stats).T

# 加入友善名稱
friendly_names = {
    "step1_law_parser": "Step 1: Law Parser",
    "step2_completion": "Step 2: Law Completion",
    "step3_json_valid": "Step 3: JSON 檢查",
    "step4_varspec": "Step 4: VarSpec 變數提取",
    "step5_constraints_parseable": "Step 5: Constraints 可解析性",
    "step5_repair_needed": "Step 5.1: 是否需要修補",
    "step5_repair_success": "Step 5.2: 修補成功",
    "step6_consistency_check": "Step 6: 一致性檢查",
    "step6_repair_success": "Step 6.2: 修補一致性",
    "step7_case_mapper": "Step 7: 案例映射",
    "step7_z3_validation": "Step 7.2: Z3 驗證",
    "step8_case_law_check": "Step 8: 法規比對",
    "step8_violation_detected": "Step 8.2: 違規偵測",
    "step8_repair_success": "Step 8.3: 修補成功",
    "step9_z3_optimize": "Step 9: Z3 最佳化",
}
checkpoint_df.rename(index=friendly_names, inplace=True)

# ============================================================
# 5️⃣ 額外表格：Performance 與 Failures
# ============================================================
performance_df = checkpoint_df.copy()
performance_df["Pass Rate (%)"] = performance_df["Pass Rate"].str.replace("%", "").astype(float)
performance_df["Avg Time (sec)"] = stats_df["total_time_sec"].mean()
performance_df = performance_df.sort_values("Pass Rate (%)")

failures_df = stats_df[~stats_df["success"]].copy()

# ============================================================
# 6️⃣ 輸出 Excel
# ============================================================
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    stats_df.to_excel(writer, sheet_name="Summary", index=False)
    summary_df.to_excel(writer, sheet_name="Overall", index=False)
    checkpoint_df.to_excel(writer, sheet_name="Checkpoints")
    performance_df.to_excel(writer, sheet_name="Performance")
    failures_df.to_excel(writer, sheet_name="Failures", index=False)

# ============================================================
# 7️⃣ 套用顏色格式（PASS/FAIL/NOT_RUN）
# ============================================================
wb = load_workbook(excel_path)
ws = wb["Summary"]

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

for col in checkpoint_cols:
    col_idx = stats_df.columns.get_loc(col) + 1
    for cell in ws.iter_rows(min_col=col_idx + 1, max_col=col_idx + 1, min_row=2):
        c = cell[0]
        if c.value == "PASS":
            c.fill = green
        elif c.value == "FAIL":
            c.fill = red
        elif c.value == "NOT_RUN":
            c.fill = gray

wb.save(excel_path)
print(f"✅ Excel 檔案已儲存至: {excel_path}")
print("👉 Sheets: Summary / Overall / Checkpoints / Performance / Failures")
